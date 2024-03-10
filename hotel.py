from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import openpyxl
import pandas as pd

# 创建 Chrome webdriver 对象
driver = webdriver.Chrome()

# 访问qunar登录页面
driver.get("https://hotel.qunar.com/")

# 填写表单，模拟登录豆瓣
log = driver.find_element(By.XPATH, '//*[@id="__headerInfo_login__"]')
log.click()
choice = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[1]/div[1]/div[2]')
choice.click()
username = driver.find_element(By.XPATH, '//*[@id="username"]')
password = driver.find_element(By.XPATH, '//*[@id="password"]')
confirm = driver.find_element(By.XPATH, '//*[@id="agreement"]')
submit_btn = driver.find_element(By.XPATH, '//*[@id="app"]/div/div[2]/div/div[1]/div[3]/div/div[3]')

username.send_keys("yourID")
password.send_keys("yourPassword")
confirm.click()
submit_btn.click()

# 等待登录成功
driver.implicitly_wait(10)
cookies = driver.get_cookies()
#print(cookies)
# 输入目的地
dst = driver.find_element(By.XPATH, '//*[@id="root"]/div/section/section[1]/div[1]/div[2]/div/div/input')
dst.clear()
dst.send_keys("济南")
# 输入景点
types = driver.find_element(By.XPATH, '//*[@id="root"]/div/section/section[1]/div[1]/div[4]/div/div/input')
types.clear()
types.send_keys("景点")
# 点击搜索
search = driver.find_element(By.XPATH, '//*[@id="root"]/div/section/section[1]/div[1]/div[5]/a')
search.click()
# 输入目的地并确认
dst2 = driver.find_element(By.XPATH, '//*[@id="root"]/div/section/div[2]/div[1]/div/div/div/input')
dst2.clear()
dst2.send_keys("济南"+Keys.ENTER)
time.sleep(2)
# 输入景点并确认
types2 = driver.find_element(By.XPATH, '//*[@id="root"]/div/section/div[2]/div[3]/div/div/div/input')
types2.clear()
types2.send_keys("景点"+Keys.ENTER)
time.sleep(2)
driver.execute_script('window.scrollBy(0, 650);')

#初始化数据表
def newExcel():
    # 新建 Excel 文档和工作表
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 插入列和设置标题
    sheet.insert_cols(1, 5)
    sheet.cell(row=1, column=1).value = '酒店名称'
    sheet.cell(row=1, column=2).value = '酒店评分'
    sheet.cell(row=1, column=3).value = '酒店地区'
    sheet.cell(row=1, column=4).value = '酒店价格'
    sheet.cell(row=1, column=5).value = '酒店评论'
    # 设置列宽
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['E'].width = 100
    # 填充空白数据
    for row in range(2, 5000):
        for col in range(1, 6):
            cell = sheet.cell(row=row, column=col)
            cell.value = " "
    workbook.save('hotel_info.xlsx')

newExcel()

# 获取酒店信息
def getHotel():
    '''not use!!'''
    next_page = driver.find_element(By.XPATH, '//*[@id="root"]/div/section/section[1]/aside[1]/div[7]/p[1]')
    temIndex = 0
    while next_page.is_enabled() and temIndex<100:
        for i in range(1, 21):
            try:
                nameX = driver.find_element(By.XPATH, '//*[@id="hotel_lst_body"]/li['+ str(i) + ']/div/div[3]/p[1]/a')
                driver.execute_script('arguments[0].click();', nameX)
                time.sleep(5)
            except NoSuchElementException:
                name = ''

# 获取每个酒店的评论
def getComment():
    # 加载 Excel 文件
    df_hotel = pd.read_excel('hotel_info.xlsx', header=0)
    # 指定要更新的行列号
    next_page = driver.find_element(By.XPATH, '//*[@id="root"]/div/section/section[1]/aside[1]/div[7]/p[1]')
    # 获取酒店总数
    hotel_numX = driver.find_element(By.XPATH, '//*[@id="root"]/div/section/section[1]/aside[1]/div[2]/span[1]')
    hotel_num = int(hotel_numX.text)
    temIndex = 0 #每页+20为了不覆盖excel
    index2= 0
    while next_page.is_enabled() and index2 < hotel_num-21:
        for i in range(1, 21):
            try:
                # 获取酒店名称、评分、地区、价格
                hotel_nameX = driver.find_element(By.XPATH, '//*[@id="hotel_lst_body"]/li['+str(i)+']/div/div[3]/p[1]/a')
                hotel_name = hotel_nameX.text
                hotel_scoreX = driver.find_element(By.XPATH, '//*[@id="hotel_lst_body"]/li['+str(i)+']/div/div[3]/p[2]/span[1]')
                hotel_score = hotel_scoreX.text.replace("分", "")
                hotel_areaX = driver.find_element(By.XPATH, '//*[@id="hotel_lst_body"]/li['+str(i)+']/div/div[3]/p[3]')
                hotel_area = hotel_areaX.text
                hotel_priceX = driver.find_element(By.XPATH, '//*[@id="hotel_lst_body"]/li['+str(i)+']/div/div[2]/p[1]/a')
                hotel_price = hotel_priceX.text.replace("¥", "").replace("起", "")
                #这些信息送入excel
                df_hotel.iloc[i+temIndex-1, 3] = hotel_price
                df_hotel.iloc[i+temIndex-1, 2] = hotel_area
                df_hotel.iloc[i+temIndex-1, 1] = hotel_score
                df_hotel.iloc[i+temIndex-1, 0] = hotel_name
                #下面开始点击每个酒店获取评论
                nameX = driver.find_element(By.XPATH, '//*[@id="hotel_lst_body"]/li['+ str(i) + ']/div/div[3]/p[1]/a')
                driver.execute_script('arguments[0].click();', nameX)
                time.sleep(2)
                #转移driver
                handles = driver.window_handles
                driver.switch_to.window(handles[-1])
                # 向下滚动 1000 像素
                driver.execute_script('window.scrollBy(0, 4000);')
                #总评论数
                numsX = driver.find_element(By.XPATH, '//*[@id="root"]/div/section[2]/section/section[2]/aside[1]/aside[3]/div/ul/li[1]/span')
                nums = int(numsX.text.replace('(', '').replace(')', ''))
                #获取按钮next
                next_button = driver.find_element(By.XPATH, '//*[@id="root"]/div/section[2]/section/section[2]/aside[1]/aside[3]/div/div/div/div[4]/div/div/div/div[3]')
                index = 0
                #开始爬取评论
                numIndex = 0
                while next_button.is_enabled() and index <= int(nums/10) and numIndex <1000:
                    for y in range(1, 11):
                        try:
                            #遍历千万评论，只为送你pd.iloc[,4]
                            commentX = driver.find_element(By.XPATH, '//*[@id="root"]/div/section[2]/section/section[2]/aside[1]/aside[3]/div/div/div/div[3]/div['+str(y)+']/div[2]/div[3]/p')
                            hotel_comment = commentX.text
                            hotel_comment= '##' + hotel_comment
                            hotel_comment = str(hotel_comment)
                            existing_value = df_hotel.iloc[i+temIndex-1, 4]
                            existing_value = str(existing_value)
                            #为了保证每次将评论追加写入excel单元格，出此↓策
                            if not pd.isna(existing_value):
                                hotel_comment = existing_value + '。 ' + hotel_comment
                            df_hotel.iloc[i+temIndex-1, 4] = hotel_comment
                            numIndex += 1
                        except NoSuchElementException:
                            #没获取到该条评论文本?加空值
                            hotel_comment = ''
                            existing_value = df_hotel.iloc[0, 4]
                            existing_value = str(existing_value)
                            if not pd.isna(existing_value):
                                hotel_comment = existing_value + '。 ' + hotel_comment
                            df_hotel.iloc[0, 4] = hotel_comment
                    index +=1
                    if(index %50 == 0):
                        print(index)
                    # 若有异常则退出循环
                    try:
                        next_button.click()
                        commentX = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, f'//*[@id="root"]/div/section[2]/section/section[2]/aside[1]/aside[3]/div/div/div/div[3]/div[10]/div[2]/div[3]/p')))
                    except:
                        break
                    next_button = driver.find_element(By.XPATH, '//*[@id="root"]/div/section[2]/section/section[2]/aside[1]/aside[3]/div/div/div/div[4]/div/div/div/div[3]')
                index2 +=1
                #关闭该页网址并返还driver
                driver.close()
                driver.switch_to.window(handles[0])
            except NoSuchElementException:
                pass
        temIndex += 20
        try:
            next_page.click()
            time.sleep(1)
            next_page = driver.find_element(By.XPATH, '//*[@id="root"]/div/section/section[1]/aside[1]/div[7]/p[1]')
            if not next_page.is_enabled() or index2 >= hotel_num-21:
                break
            WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="hotel_lst_body"]/li[1]/div/div[3]/p[1]/a')))
        except:
            break
        if temIndex%100 == 0:
            sheet_name = f'sheet{temIndex//100}'
            df_hotel.to_excel('updated2.xlsx', sheet_name=sheet_name, index=False)
    df_hotel.to_excel('updated3.xlsx', sheet_name=f'sheet{(temIndex+100)//100}', index=False)


getComment()
time.sleep(2)

driver.close()
